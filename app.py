import streamlit as st
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.utils.units import cm_to_EMU
from openpyxl.drawing.xdr import XDRPositiveSize2D
from io import BytesIO
from PIL import Image as PILImage
import math 
import os 
import json # Necesario para guardar/cargar configuraciones
import shutil # Utilidad para mover o copiar archivos

# --- Constante para la persistencia ---
LAYOUTS_FILE = "layouts_data.json"

# --- Funciones de Gestión de Archivos (Persistencia) ---

def load_layouts():
    """Carga las configuraciones de layout desde el archivo JSON."""
    if os.path.exists(LAYOUTS_FILE):
        with open(LAYOUTS_FILE, 'r') as f:
            return json.load(f)
    else:
        # Configuración por defecto si no existe el archivo
        return {
            'default': {
                'name': 'Default (2x17)',
                'area_width_cm': 9.42,
                'area_height_cm': 6.8,
                'start_row': 1,
                'start_col': 1,
                'photos_per_row': 2,
                'row_jump': 17,
                'column_spacing': 4,
                'desc_row_offset': 1,
                'desc_col_offset': 0,
            }
        }

def save_layouts(layouts):
    """Guarda las configuraciones de layout en el archivo JSON."""
    try:
        with open(LAYOUTS_FILE, 'w') as f:
            json.dump(layouts, f, indent=4)
    except Exception as e:
        st.error(f"Error al guardar la configuración en disco: {e}")

# --- Funciones de apoyo (Mantener igual) ---

def remove_extension(filename):
    """Remueve la extensión del nombre de un archivo."""
    name, ext = os.path.splitext(filename)
    return name

def redimensionar_imagen(imagen_pil, max_ancho_cm, max_alto_cm, dpi=96):
    """Redimensiona una imagen de Pillow manteniendo su relación de aspecto."""
    max_ancho_pixels = max_ancho_cm * dpi / 2.54
    max_alto_pixels = max_alto_cm / 2.54 * dpi
    ancho, alto = imagen_pil.size

    ratio_ancho = max_ancho_pixels / ancho
    ratio_alto = max_alto_pixels / alto

    if ratio_ancho < 1 or ratio_alto < 1:
        ratio = min(ratio_ancho, ratio_alto)
        nuevo_ancho = int(ancho * ratio)
        nuevo_alto = int(alto * ratio)
        imagen_redimensionada = imagen_pil.resize((nuevo_ancho, nuevo_alto), PILImage.Resampling.LANCZOS) 
        return imagen_redimensionada
    return imagen_pil

c2e = cm_to_EMU
def calcular_offset(area_cm, img_cm):
    """Calcula el offset necesario para centrar una imagen dentro de un área."""
    delta_cm = (area_cm - img_cm) / 2
    return c2e(delta_cm)

# --- Inicialización del Estado de Sesión ---

if 'layouts' not in st.session_state:
    st.session_state.layouts = load_layouts() # Cargar layouts al inicio de la sesión

if 'selected_layout_name' not in st.session_state:
    st.session_state.selected_layout_name = list(st.session_state.layouts.keys())[0] if st.session_state.layouts else 'default'

if 'descriptions_enabled' not in st.session_state:
    st.session_state.descriptions_enabled = True # Nuevo estado para habilitar/deshabilitar descripciones

# --- Interfaz de Usuario en Streamlit ---

st.title("Generador de Registro Fotográfico Personalizado 📸")
st.write("Sube tu plantilla, gestiona el layout y las fotos para generar el reporte.")

# -----------------------------
## 1. Cargar Plantilla
# -----------------------------
uploaded_excel_file = st.file_uploader("1. Sube tu plantilla de Excel (.xlsx)", type=["xlsx"])

libro = None
sheet_names = []
selected_sheet_name = None

if uploaded_excel_file:
    try:
        excel_buffer = BytesIO(uploaded_excel_file.getvalue())
        libro = load_workbook(excel_buffer)
        sheet_names = libro.sheetnames
        st.success("Plantilla de Excel cargada correctamente.")
        selected_sheet_name = st.selectbox("1a. Selecciona la hoja donde insertar las fotos:", sheet_names)
    except Exception as e:
        st.error(f"Error al cargar el archivo Excel: {e}")
        libro = None 

# ---------------------------------------------
## 2. Gestión de Configuraciones y Parámetros
# ---------------------------------------------
if libro:
    
    st.subheader("2. Gestión de Configuraciones de Layout")
    
    # ---------------------------------------------
    # Nuevo: Botón Habilitar/Deshabilitar Descripciones
    # ---------------------------------------------
    st.session_state.descriptions_enabled = st.checkbox(
        "📝 Habilitar escritura de Descripciones en Excel",
        value=st.session_state.descriptions_enabled
    )
    st.markdown("---")
    
    # Selector de Configuración
    layout_names = list(st.session_state.layouts.keys())
    selected_layout_name = st.selectbox(
        "Selecciona una configuración:",
        layout_names,
        key='layout_selector',
        index=layout_names.index(st.session_state.selected_layout_name) if st.session_state.selected_layout_name in layout_names else 0
    )
    st.session_state.selected_layout_name = selected_layout_name
    current_layout = st.session_state.layouts.get(selected_layout_name, st.session_state.layouts[list(st.session_state.layouts.keys())[0]]) # Fallback
    
    # Formulario para Editar la Configuración Seleccionada
    with st.expander(f"Editar configuración: **{current_layout['name']}**"):
        
        st.info(f"Editando: **{current_layout['name']}**")

        col1, col2 = st.columns(2)
        with col1:
            area_width_cm = st.number_input(
                "Ancho máx. imagen (cm):", min_value=1.0, value=current_layout['area_width_cm'], step=0.1, key='area_width_cm_edit'
            )
            start_row = st.number_input(
                "Fila de inicio (1-based):", min_value=1, value=current_layout['start_row'], step=1, key='start_row_edit'
            )
            photos_per_row = st.number_input(
                "Fotos por fila:", min_value=1, value=current_layout['photos_per_row'], step=1, key='photos_per_row_edit'
            )
            desc_row_offset = st.number_input(
                "Desplazamiento fila descripción:", min_value=-10, value=current_layout['desc_row_offset'], step=1, key='desc_row_offset_edit'
            )
            
        with col2:
            area_height_cm = st.number_input(
                "Altura máx. imagen (cm):", min_value=1.0, value=current_layout['area_height_cm'], step=0.1, key='area_height_cm_edit'
            )
            start_col = st.number_input(
                "Columna de inicio (1-based):", min_value=1, value=current_layout['start_col'], step=1, key='start_col_edit'
            )
            column_spacing = st.number_input(
                "Espacio h. entre fotos (cols):", min_value=1, value=current_layout['column_spacing'], step=1, key='column_spacing_edit'
            )
            row_jump = st.number_input(
                "Salto de filas entre filas de fotos:", min_value=1, value=current_layout['row_jump'], step=1, key='row_jump_edit'
            )
            desc_col_offset = st.number_input(
                "Desplazamiento col. descripción:", min_value=-10, value=current_layout['desc_col_offset'], step=1, key='desc_col_offset_edit'
            )

        col_save, col_delete = st.columns(2)
        
        with col_save:
            if st.button("💾 Aplicar y **Guardar Cambios**", key='save_current_layout'):
                st.session_state.layouts[selected_layout_name].update({
                    'area_width_cm': area_width_cm,
                    'area_height_cm': area_height_cm,
                    'start_row': start_row,
                    'start_col': start_col,
                    'photos_per_row': photos_per_row,
                    'row_jump': row_jump,
                    'column_spacing': column_spacing,
                    'desc_row_offset': desc_row_offset,
                    'desc_col_offset': desc_col_offset,
                })
                save_layouts(st.session_state.layouts) # GUARDAR EN DISCO
                st.success(f"Configuración '{current_layout['name']}' actualizada y guardada persistentemente.")
                st.rerun()

        with col_delete:
            if selected_layout_name != list(st.session_state.layouts.keys())[0]: # Prevenir eliminar el primer (default)
                if st.button(f"🗑️ Eliminar Configuración '{current_layout['name']}'", key='delete_current_layout'):
                    del st.session_state.layouts[selected_layout_name]
                    st.session_state.selected_layout_name = list(st.session_state.layouts.keys())[0]
                    save_layouts(st.session_state.layouts) # GUARDAR EN DISCO
                    st.success(f"Configuración '{current_layout['name']}' eliminada.")
                    st.rerun()
            else:
                 st.caption("No puedes eliminar la configuración por defecto.")

    st.markdown("---")
    
    # Formulario para Agregar Nueva Configuración
    with st.expander("➕ Agregar Nueva Configuración"):
        new_layout_name = st.text_input("Nombre de la nueva configuración:", value="Nuevo Layout")
        if st.button("➕ Crear Nueva Configuración", key='add_new_layout'):
            new_key = new_layout_name.strip()
            if not new_key:
                st.error("El nombre no puede estar vacío.")
            elif new_key in st.session_state.layouts:
                st.error(f"Ya existe una configuración con el nombre '{new_key}'.")
            else:
                new_layout_data = current_layout.copy()
                new_layout_data['name'] = new_layout_name
                
                st.session_state.layouts[new_key] = new_layout_data
                st.session_state.selected_layout_name = new_key
                save_layouts(st.session_state.layouts) # GUARDAR EN DISCO
                st.success(f"Configuración '{new_layout_name}' creada, guardada y seleccionada.")
                st.rerun()
                
    st.markdown("---")
    
    # --- Aplicar la configuración seleccionada a las variables de trabajo ---
    area_width_cm = current_layout['area_width_cm']
    area_height_cm = current_layout['area_height_cm']
    start_row = current_layout['start_row']
    start_col = current_layout['start_col']
    photos_per_row = current_layout['photos_per_row']
    row_jump = current_layout['row_jump']
    column_spacing = current_layout['column_spacing']
    desc_row_offset = current_layout['desc_row_offset']
    desc_col_offset = current_layout['desc_col_offset']

    # -----------------------------
    ## 3. Cargar Fotos y Descripciones
    # -----------------------------
    st.subheader("3. Cargar Registros Fotográficos (Imágenes)")
    uploaded_files = st.file_uploader("Sube tus fotos", accept_multiple_files=True,
                                        type=["png", "jpg", "jpeg"])

    if uploaded_files:
        st.write("**Vista Previa de las Fotos y Descripciones:**")

        current_file_names = [f.name for f in uploaded_files]
        if 'uploaded_file_names' not in st.session_state or \
           st.session_state.uploaded_file_names != current_file_names:
            st.session_state.rotations = {}
            st.session_state.descriptions = {}
            st.session_state.uploaded_file_names = current_file_names 

            for i, file in enumerate(uploaded_files):
                 file_id = f"file_{i}_{file.name}"
                 st.session_state.rotations[file_id] = 0
                 st.session_state.descriptions[file_id] = remove_extension(file.name)

        for i, file in enumerate(uploaded_files):
            file_id = f"file_{i}_{file.name}"
            current_rotation = st.session_state.rotations.get(file_id, 0)

            preview_col, desc_col = st.columns([1, 2])

            with preview_col:
                try:
                    img_for_preview = PILImage.open(file)
                    img_for_preview_rotated = img_for_preview.rotate(current_rotation, expand=True)
                    st.image(img_for_preview_rotated, caption=f"Foto {i+1}", width=150)

                    col_rot_left, col_rot_right = st.columns(2)
                    with col_rot_left:
                        if st.button("↺", key=f"rotar_izq_{file_id}"):
                            st.session_state.rotations[file_id] = (current_rotation - 90) % 360
                            st.rerun()
                    with col_rot_right:
                        if st.button("↻", key=f"rotar_der_{file_id}"):
                            st.session_state.rotations[file_id] = (current_rotation + 90) % 360
                            st.rerun()

                except Exception as e:
                    st.error(f"Error al cargar o mostrar {file.name}: {e}")

            with desc_col:
                 description_key = f"desc_input_{file_id}" 
                 # Solo mostramos el campo de texto si las descripciones están habilitadas, 
                 # pero guardamos el valor en session_state igual.
                 if st.session_state.descriptions_enabled:
                    st.session_state.descriptions[file_id] = st.text_input(
                        f"Descripción ({file.name}):",
                        value=st.session_state.descriptions.get(file_id, ""),
                        key=description_key
                    )
                 else:
                    st.session_state.descriptions[file_id] = st.session_state.descriptions.get(file_id, remove_extension(file.name))
                    st.caption("Descripciones deshabilitadas. No se guardará texto en Excel.")
                    

            st.markdown("---")

        # -----------------------------
        ## 4. Generar Reporte
        # -----------------------------
        st.subheader("4. Generar Reporte")
        if st.button("🚀 Generar Excel"):
            if not uploaded_excel_file:
                st.warning("Por favor, sube una plantilla de Excel.")
            elif not selected_sheet_name:
                 st.warning("Por favor, selecciona la hoja donde insertar las fotos.")
            else:
                try:
                    hoja = libro[selected_sheet_name]

                    fila_actual_base = start_row - 1
                    columna_actual_base = start_col - 1

                    for i, archivo_subido in enumerate(uploaded_files):
                        file_id = f"file_{i}_{archivo_subido.name}"

                        # Obtener ángulo de rotación y descripción
                        angulo_rotacion = st.session_state.rotations.get(file_id, 0)
                        description_text = st.session_state.descriptions.get(file_id, "")

                        # Calcular la posición de la imagen y la descripción
                        img_row_idx = fila_actual_base + math.floor(i / photos_per_row) * row_jump
                        img_col_idx = columna_actual_base + (i % photos_per_row) * column_spacing

                        desc_row_idx = img_row_idx + desc_row_offset
                        desc_col_idx = img_col_idx + desc_col_offset

                        # --- Escribir la descripción (APLICANDO EL TOGGLE) ---
                        if st.session_state.descriptions_enabled:
                            try:
                                hoja.cell(row=desc_row_idx + 1, column=desc_col_idx + 1, value=description_text)
                            except Exception as cell_error:
                                st.warning(f"No se pudo escribir la descripción para {archivo_subido.name}. Error: {cell_error}")
                        # -----------------------------------------------------

                        # Rotar y Redimensionar la imagen
                        img_pil_original = PILImage.open(archivo_subido)
                        img_pil_rotada = img_pil_original.rotate(angulo_rotacion, expand=True)
                        img_redimensionada = redimensionar_imagen(img_pil_rotada, area_width_cm, area_height_cm)

                        img_width_cm = img_redimensionada.size[0] * 2.54 / 96
                        img_height_cm = img_redimensionada.size[1] * 2.54 / 96

                        # Convertir la imagen redimensionada para openpyxl
                        img_buffer = BytesIO()
                        img_redimensionada.save(img_buffer, format="PNG")
                        img_buffer.seek(0)
                        img = Image(img_buffer)

                        # Calcular offsets para centrar y definir anclaje
                        x_offset_emu = calcular_offset(area_width_cm, img_width_cm)
                        y_offset_emu = calcular_offset(area_height_cm, img_height_cm)

                        marker = AnchorMarker(col=img_col_idx, colOff=x_offset_emu,
                                              row=img_row_idx, rowOff=y_offset_emu)

                        size = XDRPositiveSize2D(cx=c2e(img_width_cm), cy=c2e(img_height_cm))
                        img.anchor = OneCellAnchor(_from=marker, ext=size)

                        hoja.add_image(img)

                    # Guardar el libro modificado y ofrecer descarga
                    output_buffer = BytesIO()
                    libro.save(output_buffer)
                    output_buffer.seek(0)

                    base_name = remove_extension(uploaded_excel_file.name)
                    output_filename = f"{base_name}_Generado.xlsx"

                    st.download_button(
                        label="Descargar Excel Generado",
                        data=output_buffer,
                        file_name=output_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    st.success("¡El archivo Excel con las imágenes y descripciones ha sido generado exitosamente!")

                except Exception as e:
                    st.error(f"Ocurrió un error al generar el Excel: {e}")


# --- Mensaje si no se ha cargado el Excel ---
if not uploaded_excel_file:
    st.info("Por favor, carga un archivo Excel para comenzar.")